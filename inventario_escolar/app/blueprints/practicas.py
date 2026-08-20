import os
import io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required
from sqlalchemy import func
from werkzeug.utils import secure_filename
from app.extensions import db
from app.models import Alumno, Carrera, Expediente, Documento, Practica, Dependencia
from app.decorators import roles_required, active_query

practicas_bp = Blueprint('practicas', __name__)

MODULO_LABEL = 'Prácticas Profesionales'
MODULO_PREFIX = 'practicas'


@practicas_bp.before_request
@login_required
@roles_required('Super Admin', 'Practicas')
def before_request():
    pass


# ── Menú principal ──────────────────────────────────────────────────────────

@practicas_bp.route('/')
def menu():
    return render_template('practicas/menu.html',
                           modulo_label=MODULO_LABEL,
                           modulo_prefix=MODULO_PREFIX)


# ── Submódulo: Dashboard ───────────────────────────────────────────────────

@practicas_bp.route('/dashboard')
def dashboard():
    # 1. Alumnos con y sin derecho a realizar Prácticas profesionales
    # Se consideran alumnos en el semestre correspondiente (5º o superior, o egresados)
    # Para el año 2026, esto corresponde a la generación 2024 o anteriores
    current_year = datetime.now().year
    limite_generacion = (current_year - 2) % 100  # ej: 24 para 2026
    
    alumnos_elegibles = Alumno.query.filter(
        Alumno.is_deleted == False,
        Alumno.estatus.in_(['Activo', 'Egresado']),
        Alumno.anio_generacion <= limite_generacion
    ).all()
    
    tienen_derecho = 0
    no_tienen_derecho = 0
    for a in alumnos_elegibles:
        if _check_ss_completo(a):
            tienen_derecho += 1
        else:
            no_tienen_derecho += 1
            
    # 2. Datos generales de la tabla de Prácticas
    total_practicas = Practica.query.filter_by(is_deleted=False).count()
    concluidas = Practica.query.filter_by(is_deleted=False, observaciones='CONCLUIDO').count()
    en_tramite = Practica.query.filter_by(is_deleted=False, observaciones='EN TRÁMITE').count()
    sin_estatus = total_practicas - concluidas - en_tramite
    
    # 3. Distribución por Proceso
    procesos_counts = db.session.query(
        Practica.proceso, func.count(Practica.id)
    ).filter(
        Practica.is_deleted == False
    ).group_by(
        Practica.proceso
    ).all()
    
    proceso_data = {p: 0 for p in Practica.PROCESOS}
    for proc, count in procesos_counts:
        if proc in proceso_data:
            proceso_data[proc] = count
        elif proc:  # por si hay algún proceso personalizado
            proceso_data[proc] = count
            
    return render_template('practicas/dashboard.html',
                           tienen_derecho=tienen_derecho,
                           no_tienen_derecho=no_tienen_derecho,
                           total_practicas=total_practicas,
                           concluidas=concluidas,
                           en_tramite=en_tramite,
                           sin_estatus=sin_estatus,
                           proceso_data=proceso_data,
                           modulo_label=MODULO_LABEL,
                           modulo_prefix=MODULO_PREFIX)


# ── Submódulo: Detalles de Prácticas ───────────────────────────────────────

# Mapeo ordenado: (campo_db, encabezado_excel)
COLUMNAS_MAPA = [
    ('no_registro',          'No. REGISTRO'),
    ('consecutivo',          'CONSECUTIVO'),
    ('no_constancia',        'No. CONSTANCIA'),
    ('nombre',               'NOMBRE'),
    ('nombre_minusculas',    'Nombre Minúsculas'),
    ('matricula',            'MATRÍCULA'),
    ('telefono',             'TELÉFONO'),
    ('correo_estudiante',    'CORREO ESTUDIANTE'),
    ('grado_carrera',        'GRADO-CARRERA'),
    ('turno',                'TURNO'),
    ('carrera',              'CARRERA'),
    ('observaciones',        'OBSERVACIONES'),
    ('proceso',              'PROCESO'),
    ('empresa',              'EMPRESA'),
    ('sector',               'SECTOR'),
    ('nombre_proyecto',      'NOMBRE DEL PROYECTO'),
    ('tel_empresa',          'TEL. EMPRESA'),
    ('direccion_empresa',    'DIRECCIÓN EMPRESA'),
    ('correo_empresa',       'CORR. EM'),
    ('generacion',           'GENERACIÓN'),
    ('sexo',                 'H/M'),
    ('becados',              'BECADOS'),
    ('monto',                'MONTO'),
    ('s_prac',               'S. PRÁC'),
    ('s_prac_fecha_excel',   'S. PRÁC.'),
    ('f_inicio',             'F. INICIO'),
    ('f_inicio_dia',         'DIA'),
    ('f_inicio_mes',         'MES'),
    ('f_inicio_anio',        'AÑO'),
    ('c_pres',               'C.PRES.'),
    ('c_pres_fecha_excel',   'C.PRES.2'),
    ('f_cp',                 'F.C.P'),
    ('c_acep',               'C. ACEP'),
    ('c_acep_fecha_excel',   'C. ACEP.'),
    ('f_ca',                 'F.C.A'),
    ('f_ca_dia',             'DIA 3'),
    ('f_ca_mes',             'MES4'),
    ('f_ca_anio',            'AÑO 5'),
    ('p_trabj',              'P. TRABJ'),
    ('p_trabj_fecha_excel',  'P. TRABJ.'),
    ('f_ptr',                'F. P.TR'),
    ('i_inter',              'I. INTER'),
    ('i_inter_fecha_excel',  'I. INTER.'),
    ('f_ii',                 'F.I.I.'),
    ('f_l_ii',               'F.L. I.I.'),
    ('estado',               'ESTADO'),
    ('inf_final',            'INF. FINAL'),
    ('inf_final_valor',      'INF. FINAL6'),
    ('f_i_final',            'F.I.FINAL'),
    ('estado_inf_final',     'ESTADO7'),
    ('r_e_final',            'R-E-FINAL'),
    ('r_e_final_valor',      'R-E-FINAL8'),
    ('f_re_final',           'F.R-E FINAL'),
    ('cons_t',               'CONS.T'),
    ('c_t_valor',            'C.T'),
    ('f_ct',                 'F.C.T'),
    ('f_ct_dia',             'DIA2'),
    ('f_ct_mes',             'MES2'),
    ('f_ct_anio',            'AÑO2'),
    ('promedio',             'PROMEDIO'),
    ('tiempo_proceso',       'TIEMPO DEL PROCESO'),
    ('resumen_observaciones','RESUMEN -OBSERVACIONES'),
    ('carpeta',              'CARPETA'),
    ('paso_por_constancia',  'PASO POR CONSTANCIA'),
    ('pc',                   'P.C.'),
]

# Filtros selector: (query_param_name, campo_db, opciones)
FILTROS_SELECTOR = [
    ('f_grado_carrera',   'grado_carrera',    Practica.GRADOS_CARRERA,    'Grado-Carrera'),
    ('f_turno',           'turno',            Practica.TURNOS,            'Turno'),
    ('f_carrera',         'carrera',          Practica.CARRERAS,          'Carrera'),
    ('f_observaciones',   'observaciones',    Practica.OBSERVACIONES_OPTS,'Observaciones'),
    ('f_proceso',         'proceso',          Practica.PROCESOS,          'Proceso'),
    ('f_sector',          'sector',           Practica.SECTORES,          'Sector'),
    ('f_sexo',            'sexo',             Practica.SEXOS,             'H/M'),
    ('f_s_prac',          's_prac',           Practica.OPTS_SNC,          'S. PRÁC'),
    ('f_c_pres',          'c_pres',           Practica.OPTS_SNC,          'C.PRES.'),
    ('f_c_acep',          'c_acep',           Practica.OPTS_SNC,          'C. ACEP'),
    ('f_p_trabj',         'p_trabj',          Practica.OPTS_SNC,          'P. TRABJ'),
    ('f_i_inter',         'i_inter',          Practica.OPTS_SNC,          'I. INTER'),
    ('f_estado',          'estado',           Practica.ESTADOS,           'Estado (I.I.)'),
    ('f_inf_final',       'inf_final',        Practica.OPTS_SNC,          'INF. FINAL'),
    ('f_estado_inf',      'estado_inf_final', Practica.ESTADOS,           'Estado (Inf. Final)'),
    ('f_r_e_final',       'r_e_final',        Practica.OPTS_SNC,          'R-E-FINAL'),
    ('f_cons_t',          'cons_t',           Practica.OPTS_SNC,          'CONS.T'),
]


def _build_detalles_query():
    """Construye la query filtrada de Practica según request.args."""
    query = Practica.query.filter_by(is_deleted=False)

    # Búsqueda general por texto
    search = request.args.get('search', '').strip()
    if search:
        query = query.filter(
            (Practica.nombre.ilike(f'%{search}%')) |
            (Practica.matricula.ilike(f'%{search}%')) |
            (Practica.empresa.ilike(f'%{search}%')) |
            (Practica.no_constancia.ilike(f'%{search}%'))
        )

    # Filtros de tipo selector
    for param_name, db_field, _options, _label in FILTROS_SELECTOR:
        val = request.args.get(param_name)
        if val:
            query = query.filter(getattr(Practica, db_field) == val)

    # Filtro por generación (texto libre)
    gen = request.args.get('f_generacion', '').strip()
    if gen:
        query = query.filter(Practica.generacion.ilike(f'%{gen}%'))

    return query.order_by(Practica.no_registro)


def _get_columnas_activas():
    """Obtiene las columnas visibles de request.args o retorna todas."""
    cols = request.args.getlist('cols')
    all_fields = [c[0] for c in COLUMNAS_MAPA]
    if not cols:
        return all_fields
    # Validar contra campos existentes
    return [c for c in cols if c in all_fields]


@practicas_bp.route('/detalles')
def detalles():
    query = _build_detalles_query()
    practicas = query.all()
    columnas_activas = _get_columnas_activas()

    # Calcular cuántos filtros están activos para el badge
    filtros_activos = 0
    if request.args.get('search', '').strip():
        filtros_activos += 1
    if request.args.get('f_generacion', '').strip():
        filtros_activos += 1
    for param_name, _db, _opts, _label in FILTROS_SELECTOR:
        if request.args.get(param_name):
            filtros_activos += 1

    return render_template('practicas/detalles.html',
                           practicas=practicas,
                           columnas_mapa=COLUMNAS_MAPA,
                           columnas_activas=columnas_activas,
                           filtros_selector=FILTROS_SELECTOR,
                           filtros_activos=filtros_activos,
                           modulo_label=MODULO_LABEL,
                           modulo_prefix=MODULO_PREFIX)


@practicas_bp.route('/detalles/exportar')
def exportar_detalles():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = _build_detalles_query()
    practicas = query.all()
    columnas_activas = _get_columnas_activas()

    # Filtrar solo las columnas visibles
    cols_export = [(f, h) for f, h in COLUMNAS_MAPA if f in columnas_activas]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Prácticas Profesionales'

    # Estilos para el encabezado
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Escribir encabezados
    for col_idx, (field, header) in enumerate(cols_export, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Escribir datos
    for row_idx, p in enumerate(practicas, start=2):
        for col_idx, (field, _header) in enumerate(cols_export, start=1):
            val = getattr(p, field, None)
            # Formatear fechas para Excel
            if val is not None and hasattr(val, 'strftime'):
                val = val.strftime('%d/%m/%Y')
            # Convertir Decimal a float
            if val is not None and hasattr(val, 'is_finite'):
                val = float(val)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Autoajustar ancho de columnas
    for col_idx, (field, header) in enumerate(cols_export, start=1):
        max_len = len(header)
        for row in range(2, min(len(practicas) + 2, 52)):  # Sample first 50 rows
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 45)

    # Congelar la primera fila
    ws.freeze_panes = 'A2'

    # Guardar en buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Practicas_Profesionales_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ── Submódulo: Alumnos ─────────────────────────────────────────────────────

def _check_ss_completo(alumno):
    """
    Verifica si el alumno completó con éxito su Servicio Social.
    Retorna True si tiene un expediente SS (tipo_modulo='s') con al menos
    un documento y TODOS en estado 'Entregado'.
    """
    expediente_ss = Expediente.query.filter_by(
        alumno_id=alumno.id,
        tipo_modulo='s',
        is_deleted=False
    ).first()
    if not expediente_ss:
        return False
    total_docs = Documento.query.filter_by(
        expediente_id=expediente_ss.id,
        is_deleted=False
    ).count()
    if total_docs == 0:
        return False
    entregados = Documento.query.filter_by(
        expediente_id=expediente_ss.id,
        is_deleted=False,
        estado='Entregado'
    ).count()
    return entregados == total_docs


def _get_practica_for_alumno(alumno):
    """Busca un registro de prácticas existente para el alumno por matrícula."""
    if not alumno.matricula:
        return None
    return Practica.query.filter_by(
        matricula=alumno.matricula,
        is_deleted=False
    ).first()


@practicas_bp.route('/alumnos')
def alumnos():
    page = request.args.get('page', 1, type=int)
    carrera_filter = request.args.get('carrera_filter')
    search = request.args.get('search')
    estatus_filter = request.args.get('estatus_filter')
    solo_aptos = request.args.get('solo_aptos')

    query = active_query(Alumno)

    if carrera_filter:
        query = query.filter(Alumno.carrera_id == int(carrera_filter))
    if search:
        query = query.filter(
            (Alumno.nombre.ilike(f'%{search}%')) |
            (Alumno.matricula.ilike(f'%{search}%'))
        )
    if estatus_filter:
        query = query.filter(Alumno.estatus == estatus_filter)

    query = query.order_by(Alumno.nombre)
    pagination = query.paginate(page=page, per_page=20, error_out=False)

    carreras = active_query(Carrera).all()
    estatuses = ['Activo', 'Inactivo', 'Egresado']

    # Build enriched data for each alumno
    alumnos_data = []
    for alumno in pagination.items:
        ss_completo = _check_ss_completo(alumno)
        practica = _get_practica_for_alumno(alumno)
        # If "solo_aptos" filter is on, skip non-apto alumnos
        # (we'll filter after pagination for simplicity — the filter is cosmetic)
        alumnos_data.append({
            'alumno': alumno,
            'ss_completo': ss_completo,
            'apto_practicas': ss_completo,
            'tiene_practica': practica is not None,
            'practica_id': practica.id if practica else None,
        })

    # Apply solo_aptos filter post-query if needed
    if solo_aptos:
        alumnos_data = [a for a in alumnos_data if a['apto_practicas']]

    return render_template('practicas/alumnos.html',
                           alumnos_data=alumnos_data,
                           pagination=pagination,
                           carreras=carreras,
                           estatuses=estatuses,
                           carrera_filter=carrera_filter,
                           search=search,
                           estatus_filter=estatus_filter,
                           solo_aptos=solo_aptos,
                           modulo_label=MODULO_LABEL,
                           modulo_prefix=MODULO_PREFIX)


# ── Asignar Prácticas (formulario) ─────────────────────────────────────────

@practicas_bp.route('/alumnos/<int:alumno_id>/asignar', methods=['GET', 'POST'])
def asignar(alumno_id):
    alumno = active_query(Alumno).filter_by(id=alumno_id).first_or_404()

    if request.method == 'POST':
        practica = Practica()

        # Helper to get optional form values
        def fv(name):
            val = request.form.get(name)
            return val if val and val.strip() else None

        def fv_int(name):
            val = fv(name)
            if val:
                try:
                    return int(val)
                except (ValueError, TypeError):
                    return None
            return None

        def fv_float(name):
            val = fv(name)
            if val:
                try:
                    return float(val.replace(',', '.'))
                except (ValueError, TypeError):
                    return None
            return None

        def fv_date(name):
            val = fv(name)
            if val:
                try:
                    return datetime.strptime(val, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    return None
            return None

        # ── Col 1-8: Datos personales ──────────────────────────────────
        practica.no_registro = fv_int('no_registro')
        practica.consecutivo = fv_int('consecutivo')
        practica.no_constancia = fv('no_constancia')
        practica.nombre = fv('nombre')
        practica.nombre_minusculas = fv('nombre_minusculas')
        practica.matricula = fv('matricula')
        practica.telefono = fv('telefono')
        practica.correo_estudiante = fv('correo_estudiante')

        # ── Col 9-13: Académicos ───────────────────────────────────────
        practica.grado_carrera = fv('grado_carrera')
        practica.turno = fv('turno')
        practica.carrera = fv('carrera')
        practica.observaciones = fv('observaciones')
        practica.proceso = fv('proceso')

        # ── Col 14-19: Empresa ─────────────────────────────────────────
        practica.empresa = fv('empresa')
        practica.sector = fv('sector')
        practica.nombre_proyecto = fv('nombre_proyecto')
        practica.tel_empresa = fv('tel_empresa')
        practica.direccion_empresa = fv('direccion_empresa')
        practica.correo_empresa = fv('correo_empresa')

        # ── Col 20-23: Personales ──────────────────────────────────────
        practica.generacion = fv('generacion')
        practica.sexo = fv('sexo')
        practica.becados = fv_float('becados')
        practica.monto = fv('monto')

        # ── Col 24-29: Solicitud de Prácticas ──────────────────────────
        practica.s_prac = fv('s_prac')
        practica.s_prac_fecha_excel = fv_date('s_prac_fecha_excel')
        practica.f_inicio = fv_date('f_inicio')
        practica.f_inicio_dia = fv_int('f_inicio_dia')
        practica.f_inicio_mes = fv('f_inicio_mes')
        practica.f_inicio_anio = fv_int('f_inicio_anio')

        # ── Col 30-32: Carta de Presentación ───────────────────────────
        practica.c_pres = fv('c_pres')
        practica.c_pres_fecha_excel = fv_date('c_pres_fecha_excel')
        practica.f_cp = fv_date('f_cp')

        # ── Col 33-38: Carta de Aceptación ─────────────────────────────
        practica.c_acep = fv('c_acep')
        practica.c_acep_fecha_excel = fv_date('c_acep_fecha_excel')
        practica.f_ca = fv_date('f_ca')
        practica.f_ca_dia = fv_int('f_ca_dia')
        practica.f_ca_mes = fv('f_ca_mes')
        practica.f_ca_anio = fv_int('f_ca_anio')

        # ── Col 39-41: Plan de Trabajo ─────────────────────────────────
        practica.p_trabj = fv('p_trabj')
        practica.p_trabj_fecha_excel = fv_date('p_trabj_fecha_excel')
        practica.f_ptr = fv_date('f_ptr')

        # ── Col 42-46: Informe Intermedio ──────────────────────────────
        practica.i_inter = fv('i_inter')
        practica.i_inter_fecha_excel = fv_date('i_inter_fecha_excel')
        practica.f_ii = fv_date('f_ii')
        practica.f_l_ii = fv_date('f_l_ii')
        practica.estado = fv('estado')

        # ── Col 47-50: Informe Final ───────────────────────────────────
        practica.inf_final = fv('inf_final')
        practica.inf_final_valor = fv_int('inf_final_valor')
        practica.f_i_final = fv_date('f_i_final')
        practica.estado_inf_final = fv('estado_inf_final')

        # ── Col 51-53: Revisión-Entrega Final ──────────────────────────
        practica.r_e_final = fv('r_e_final')
        practica.r_e_final_valor = fv_float('r_e_final_valor')
        practica.f_re_final = fv_date('f_re_final')

        # ── Col 54-59: Constancia de Terminación ───────────────────────
        practica.cons_t = fv('cons_t')
        practica.c_t_valor = fv_float('c_t_valor')
        practica.f_ct = fv_date('f_ct')
        practica.f_ct_dia = fv_int('f_ct_dia')
        practica.f_ct_mes = fv('f_ct_mes')
        practica.f_ct_anio = fv_int('f_ct_anio')

        # ── Col 60-65: Cierre ──────────────────────────────────────────
        practica.promedio = fv('promedio')
        practica.tiempo_proceso = fv('tiempo_proceso')
        practica.resumen_observaciones = fv('resumen_observaciones')
        practica.carpeta = fv('carpeta')
        practica.paso_por_constancia = fv('paso_por_constancia')
        practica.pc = fv('pc')

        db.session.add(practica)
        db.session.commit()
        flash('Registro de prácticas creado exitosamente.', 'success')
        return redirect(url_for('practicas.alumnos'))

    # GET: pre-fill from alumno data
    prefill = {
        'nombre': alumno.nombre or '',
        'nombre_minusculas': (alumno.nombre or '').title(),
        'matricula': alumno.matricula or '',
        'carrera': alumno.carrera.nombre if alumno.carrera else '',
        'generacion': alumno.generacion_completa if alumno.generacion_completa != 'Pendiente' else '',
    }

    # Fetch dependencias that are suitable for Prácticas
    dependencias = Dependencia.query.filter(
        Dependencia.tipo.in_(['Practicas', 'Ambos']),
        Dependencia.is_deleted == False
    ).order_by(Dependencia.nombre).all()

    return render_template('practicas/asignar_form.html',
                           alumno=alumno,
                           prefill=prefill,
                           practica_model=Practica,
                           dependencias=dependencias,
                           modulo_label=MODULO_LABEL,
                           modulo_prefix=MODULO_PREFIX)


# ── Ver Detalles de Práctica ───────────────────────────────────────────────

@practicas_bp.route('/alumnos/<int:alumno_id>/detalle')
def detalle_practica(alumno_id):
    alumno = active_query(Alumno).filter_by(id=alumno_id).first_or_404()
    practica = Practica.query.filter_by(
        matricula=alumno.matricula,
        is_deleted=False
    ).first_or_404()

    return render_template('practicas/detalle_practica.html',
                           alumno=alumno,
                           practica=practica,
                           modulo_label=MODULO_LABEL,
                           modulo_prefix=MODULO_PREFIX)


# ── Submódulo: Importación Masiva ──────────────────────────────────────────

@practicas_bp.route('/importar', methods=['GET', 'POST'])
def importar():
    resultado = None
    if request.method == 'POST':
        if 'archivo_excel' not in request.files:
            flash('No se subió ningún archivo.', 'danger')
            return redirect(request.url)

        file = request.files['archivo_excel']
        if file.filename == '':
            flash('No se seleccionó ningún archivo.', 'danger')
            return redirect(request.url)

        if file and file.filename.endswith(('.xlsx', '.xls')):
            # Ensure the /tmp or similar directory exists, or just use a local tmp folder
            import tempfile
            temp_dir = tempfile.gettempdir()
            filepath = os.path.join(temp_dir, secure_filename(file.filename))
            file.save(filepath)
            
            resultado = procesar_excel_practicas(filepath)
            os.remove(filepath)
            
            if resultado['errores']:
                flash('Importación completada con algunos errores. Revisa el resumen.', 'warning')
            else:
                flash('Importación de prácticas completada exitosamente.', 'success')
        else:
            flash('Formato de archivo no válido. Usa .xlsx o .xls', 'danger')

    return render_template('practicas/importar.html', 
                           resultado=resultado,
                           columnas_mapa=COLUMNAS_MAPA,
                           modulo_label=MODULO_LABEL,
                           modulo_prefix=MODULO_PREFIX)


def procesar_excel_practicas(filepath):
    import pandas as pd
    import math

    try:
        df = pd.read_excel(filepath, engine='openpyxl')
    except Exception as e:
        return {'insertados': 0, 'duplicados': 0, 'errores': [f'Error al leer el archivo: {str(e)}']}

    # Clean headers to match the COLUMNAS_MAPA precisely
    df.columns = [str(c).strip() for c in df.columns]

    # Create a reverse map: excel_header -> db_field
    header_to_field = {header.strip(): field for field, header in COLUMNAS_MAPA}

    # Verify if we have at least 'MATRÍCULA' and 'No. CONSTANCIA'
    if 'MATRÍCULA' not in df.columns or 'No. CONSTANCIA' not in df.columns:
        return {'insertados': 0, 'duplicados': 0, 'errores': ['Faltan columnas clave: MATRÍCULA y/o No. CONSTANCIA']}

    insertados = 0
    duplicados = 0
    errores = []

    # Get types for DB fields to cast properly
    date_fields = [f for f in dir(Practica) if getattr(getattr(Practica, f), 'type', None) and str(getattr(Practica, f).type).startswith('DATE')]
    numeric_fields = ['becados', 'r_e_final_valor', 'c_t_valor']
    integer_fields = ['no_registro', 'consecutivo', 'f_inicio_dia', 'f_inicio_anio', 'f_ca_dia', 'f_ca_anio', 'inf_final_valor', 'f_ct_dia', 'f_ct_anio']

    for idx, row in df.iterrows():
        fila_num = idx + 2
        try:
            matricula = str(row['MATRÍCULA']).strip()
            no_constancia = str(row['No. CONSTANCIA']).strip()

            if (matricula == 'nan' or not matricula) and (no_constancia == 'nan' or not no_constancia):
                # Skip totally empty rows
                continue

            # Check if exists
            existente = Practica.query.filter(
                (Practica.matricula == matricula) & (Practica.no_constancia == no_constancia),
                Practica.is_deleted == False
            ).first()

            if existente:
                duplicados += 1
                continue

            nueva_practica = Practica()

            for col in df.columns:
                if col in header_to_field:
                    field = header_to_field[col]
                    val = row[col]

                    if pd.isna(val):
                        continue
                    
                    if field in date_fields:
                        if isinstance(val, pd.Timestamp):
                            setattr(nueva_practica, field, val.date())
                        else:
                            # Try to parse if it's a string, or leave None
                            pass
                    elif field in numeric_fields:
                        try:
                            setattr(nueva_practica, field, float(val))
                        except (ValueError, TypeError):
                            pass
                    elif field in integer_fields:
                        try:
                            setattr(nueva_practica, field, int(float(val)))
                        except (ValueError, TypeError):
                            pass
                    else:
                        val_str = str(val).strip()
                        if val_str != 'nan':
                            # Truncate strings to prevent DB overflow if needed, but for now just assign
                            setattr(nueva_practica, field, val_str)

            db.session.add(nueva_practica)
            insertados += 1

        except Exception as e:
            errores.append(f'Fila {fila_num}: {str(e)}')

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        errores.append(f'Error al guardar en base de datos: {str(e)}')

    return {'insertados': insertados, 'duplicados': duplicados, 'errores': errores}
